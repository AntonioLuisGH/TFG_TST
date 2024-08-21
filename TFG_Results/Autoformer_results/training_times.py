import os
import pandas as pd
import re
from openpyxl import load_workbook


def extract_number(folder_name):
    match = re.search(r'\d+', folder_name)
    return int(match.group()) if match else float('inf')


# Get the path to the main folder (the same where the script is located)
main_folder = os.getcwd()

# Initialize a list to store the data
data = []

# Iterate over all subfolders in the main folder and sort them
for folder_name in sorted(os.listdir(main_folder), key=extract_number):
    folder_path = os.path.join(main_folder, folder_name)
    if os.path.isdir(folder_path) and folder_name.startswith('A'):
        # Path to the plots subfolder within the current folder
        plots_folder = os.path.join(folder_path, 'plots')
        # Path to the metrics.txt file in the plots subfolder
        metrics_file = os.path.join(plots_folder, 'metrics.txt')
        # Path to the training_time.txt file in the plots subfolder
        training_time_file = os.path.join(plots_folder, 'training_time.txt')

        training_time = None
        if os.path.exists(training_time_file):
            # Read the training_time.txt file and extract the training time
            with open(training_time_file, 'r') as file:
                # Example format: "Training time for 'Transformer_Loss': 275.11 seconds"
                line = file.readline().strip()
                match = re.search(r'(\d+\.\d+|\d+) seconds', line)
                if match:
                    training_time = float(match.group(1))

        if os.path.exists(metrics_file):
            # Read the metrics.txt file and extract the data
            with open(metrics_file, 'r') as file:
                lines = file.readlines()
                # Skip the first header line
                for line in lines[1:]:
                    # Split the line into components
                    components = line.split()
                    if len(components) == 3:
                        variable, mse, r_squared = components
                        # Add a row to the data
                        data.append([folder_name, variable, float(
                            mse), float(r_squared), training_time])

# Create a pandas DataFrame with the data
df = pd.DataFrame(data, columns=['Folder', 'Variable', 'MSE', 'R_squared', 'Training_time'])

# Extraer el número de las carpetas y crear una nueva columna numérica para ordenar
df['Folder_num'] = df['Folder'].str.extract('(\d+)').astype(int)

# Pivotar el DataFrame para que las filas sean las carpetas y las columnas sean las variables para MSE
df_pivoted_mse = df.pivot_table(
    index='Folder', columns='Variable', values='MSE', aggfunc='first')
df_pivoted_mse = df_pivoted_mse.reindex(df.groupby(
    'Folder')['Folder_num'].first().sort_values().index)

# Pivotar el DataFrame para que las filas sean las carpetas y las columnas sean las variables para R_squared
df_pivoted_r_squared = df.pivot_table(
    index='Folder', columns='Variable', values='R_squared', aggfunc='first')
df_pivoted_r_squared = df_pivoted_r_squared.reindex(
    df.groupby('Folder')['Folder_num'].first().sort_values().index)

# Pivotar el DataFrame para que las filas sean las carpetas y la columna sea el tiempo de entrenamiento
df_pivoted_training_time = df.pivot_table(
    index='Folder', values='Training_time', aggfunc='first')
df_pivoted_training_time = df_pivoted_training_time.reindex(
    df.groupby('Folder')['Folder_num'].first().sort_values().index)

# Guardar ambos DataFrames en un archivo Excel con tres hojas
output_file = 'output.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_pivoted_mse.to_excel(writer, sheet_name='MSE')
    df_pivoted_r_squared.to_excel(writer, sheet_name='R_squared')
    df_pivoted_training_time.to_excel(writer, sheet_name='Training_Time')

# Cargar el archivo Excel y ajustar el tamaño de las celdas
wb = load_workbook(output_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Ajustar la altura de las filas
    for row in ws.iter_rows():
        # Aquí puedes cambiar el valor para ajustar la altura de las filas
        ws.row_dimensions[row[0].row].height = 20

# Guardar los cambios
wb.save(output_file)

print("Archivo Excel creado con éxito y tamaño de celdas ajustado.")
